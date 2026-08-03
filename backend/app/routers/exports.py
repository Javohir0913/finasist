"""Выгрузка отчётов и реестров в Excel (.xlsx) — как листы исходной книги.

Каждый отчёт описан набором колонок; данные берутся из тех же функций, что и
JSON-эндпоинты, поэтому цифры в файле и на экране всегда совпадают.
"""
from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from ..database import get_db
from ..ledger import LEDGERS
from ..models import (
    Employee,
    MaterialIssue,
    MaterialReceipt,
    PayrollEntry,
    Production,
    Sale,
    Service,
    Transaction,
    User,
)
from ..periods import history_guard
from ..security import require
from . import reports as R
from .payroll import payroll_summary

router = APIRouter(
    prefix="/api/export",
    tags=["export"],
    dependencies=[Depends(history_guard)],
)

MONTHS = ["", "январь", "февраль", "март", "апрель", "май", "июнь",
          "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]

HEAD_FILL = PatternFill("solid", fgColor="1F3A5F")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13)
TOTAL_FONT = Font(bold=True)
THIN = Side(style="thin", color="BFC9D4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = "#,##0.00"


def _period_title(year: int | None, month: int | None) -> str:
    if year and month:
        return f"за {MONTHS[int(month)]} {year} года"
    if year:
        return f"за {year} год"
    return "за весь период"


def _sheet(wb: Workbook, title: str, subtitle: str, columns: list[dict]) -> None:
    """Создать лист с шапкой; columns: [{key, label, width?, money?}]."""
    ws = wb.create_sheet(title[:31])
    ws.cell(1, 1, subtitle).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))
    for i, col in enumerate(columns, 1):
        c = ws.cell(3, i, col["label"])
        c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = col.get("width", 16)
    ws.freeze_panes = "A4"


def _fill(ws, columns: list[dict], rows: list[dict], start: int = 4) -> int:
    r = start
    for row in rows:
        for i, col in enumerate(columns, 1):
            value = row.get(col["key"], "")
            c = ws.cell(r, i, value)
            c.border = BORDER
            if col.get("money"):
                c.number_format = MONEY
        r += 1
    return r


def _totals(ws, columns: list[dict], row_idx: int, totals: dict, label: str = "ИТОГО") -> None:
    ws.cell(row_idx, 1, label).font = TOTAL_FONT
    for i, col in enumerate(columns, 1):
        if col["key"] in totals:
            c = ws.cell(row_idx, i, totals[col["key"]])
            c.font, c.number_format, c.border = TOTAL_FONT, MONEY, BORDER


def _xlsx(wb: Workbook, filename: str) -> Response:
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    if not wb.sheetnames:  # книга без листов не сохраняется
        wb.create_sheet("Нет данных")
    if wb.sheetnames == ["Sheet"]:
        wb["Sheet"].title = "Нет данных"
        wb["Нет данных"].cell(1, 1, "За выбранный период данных нет").font = TITLE_FONT
    buf = BytesIO()
    wb.save(buf)
    # заголовки HTTP — latin-1, поэтому кириллица уходит через RFC 5987
    ascii_name = filename.encode("ascii", "ignore").decode() or "report.xlsx"
    quoted = quote(filename)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
                f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{quoted}',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


def cols(*spec) -> list[dict]:
    """cols(("key","Заголовок",width,money), ...)"""
    out = []
    for key, label, *rest in spec:
        width = rest[0] if rest else 16
        money = rest[1] if len(rest) > 1 else False
        out.append({"key": key, "label": label, "width": width, "money": money})
    return out


# ============================ ОТЧЁТЫ ============================
LEDGER_COLS = cols(
    ("inn", "ИНН", 14), ("name", "Наименование организации", 42),
    ("open_debit", "На начало: Дебет", 18, True),
    ("open_credit", "На начало: Кредит", 18, True),
    ("turn_debit", "Оборот: Дебет", 18, True),
    ("turn_credit", "Оборот: Кредит", 18, True),
    ("end_debit", "На конец: Дебет", 18, True),
    ("end_credit", "На конец: Кредит", 18, True),
)

EXPENSE_COLS = cols(
    ("code", "Код платежа", 14), ("name", "НАИМЕНОВАНИЕ", 46), ("qty", "Кол-во", 10),
    ("bank_uzs", "БАНК UZS", 18, True), ("bank_usd", "БАНК USD", 18, True),
    ("kassa_uzs", "КАССА UZS", 18, True), ("kassa_usd", "КАССА USD", 18, True),
    ("accrued_uzs", "Начислено", 18, True),
    ("total_uzs", "ВСЕГО UZS", 18, True), ("total_usd", "ВСЕГО USD", 18, True),
)

CASHFLOW_COLS = cols(
    ("code", "Код платежа", 14), ("name", "НАИМЕНОВАНИЕ", 46),
    ("bank_in", "БАНК приход", 18, True), ("bank_out", "БАНК расход", 18, True),
    ("kassa_in", "КАССА приход", 18, True), ("kassa_out", "КАССА расход", 18, True),
    ("in", "ВСЕГО приход", 18, True), ("out", "ВСЕГО расход", 18, True),
)


async def _report_sheets(wb: Workbook, db: AsyncSession, year, month, which: set[str]):
    title = _period_title(year, month)

    if "pnl" in which:
        data = await R.pnl(year=year, month=month, division=None, _=None, db=db)
        c = cols(("code", "Код", 8), ("name", "Наименование показателя", 56),
                 ("amount", "Сумма, сум", 22, True))
        _sheet(wb, "ОФР Форма №2", f"ОТЧЁТ О ФИНАНСОВЫХ РЕЗУЛЬТАТАХ — Форма №2 {title}", c)
        spec = [
            ("010", "Чистая выручка от реализации", "revenue"),
            ("020", "Себестоимость реализованной продукции", "cogs"),
            ("030", "Валовая прибыль (010 − 020)", "gross"),
            ("040", "Расходы периода, всего", "period"),
            ("050", "   Расходы по реализации", "sell"),
            ("060", "   Административные расходы", "admin"),
            ("070", "   Прочие операционные расходы", "other"),
            ("090", "Прочие доходы", "other_income"),
            ("100", "Прибыль от основной деятельности", "op_profit"),
            ("120", "Доходы по финансовой деятельности", "fin_income"),
            ("", "   в т.ч. доходы от валютных курсовых разниц", "fx_income"),
            ("130", "Расходы по финансовой деятельности", "fin_loss"),
            ("", "   в т.ч. убытки от валютных курсовых разниц", "fx_loss"),
            ("220", "Прибыль от общехозяйственной деятельности", "gh_profit"),
            ("230", "Чрезвычайные прибыли и убытки", "extraordinary"),
            ("240", "Прибыль до уплаты налога на прибыль", "before_tax"),
            ("250", "Налог на прибыль", "tax"),
            ("260", "Прочие налоги и сборы от прибыли", "other_taxes"),
            ("270", "ЧИСТАЯ ПРИБЫЛЬ", "net"),
        ]
        _fill(wb["ОФР Форма №2"], c,
              [{"code": k, "name": n, "amount": data[key]} for k, n, key in spec])

        # свод по подразделениям — как строки Мачстон/Жби/Турк книги
        divs = await R.pnl_by_divisions(year=year, month=month, _=None, db=db)
        if divs["rows"]:
            dc = cols(("division", "Подразделение", 22), ("revenue", "Выручка", 20, True),
                      ("cogs", "Себестоимость", 20, True), ("gross", "Валовая прибыль", 20, True),
                      ("sell", "Реализация", 18, True), ("admin", "Административные", 20, True),
                      ("other", "Прочие операционные", 20, True),
                      ("period", "Расходы периода", 20, True),
                      ("op_profit", "Прибыль от основной", 22, True))
            _sheet(wb, "ОФР по подразделениям", f"ОФР в разрезе подразделений {title}", dc)
            ws = wb["ОФР по подразделениям"]
            r = _fill(ws, dc, divs["rows"])
            _totals(ws, dc, r, divs["total"], "ВСЕГО")

    if "balance" in which:
        data = await R.balance(year=year, month=month, _=None, db=db)
        c = cols(("code", "№ стр", 8), ("name", "Наименование", 60),
                 ("opening", "На начало периода", 22, True),
                 ("amount", "На конец периода", 22, True))
        _sheet(wb, "Баланс Форма №1", f"БАЛАНС — Форма №1 {title}", c)
        ws = wb["Баланс Форма №1"]

        def prep(rows_):
            return [
                {**x, "name": "   " * x["level"] + x["name"],
                 "opening": x["opening"] if x["opening"] is not None else "",
                 "amount": x["amount"] if x["amount"] is not None else ""}
                for x in rows_
            ]

        r = _fill(ws, c, prep(data["asset_rows"]))
        _fill(ws, c, prep(data["liability_rows"]), r + 1)

    if "cashflow" in which:
        data = await R.cashflow(year=year, month=month, division=None, _=None, db=db)
        _sheet(wb, "CASH FLOW", f"ДВИЖЕНИЕ ДЕНЕЖНЫХ СРЕДСТВ {title}", CASHFLOW_COLS)
        ws = wb["CASH FLOW"]
        head = [
            {"code": "", "name": "ОСТАТОК на начало периода",
             "bank_in": data["bank"]["open"], "kassa_in": data["kassa"]["open"],
             "in": data["total"]["open"]},
        ]
        r = _fill(ws, CASHFLOW_COLS, head + data["by_code"])
        _totals(ws, CASHFLOW_COLS, r, {
            "bank_in": data["bank"]["in"], "bank_out": data["bank"]["out"],
            "kassa_in": data["kassa"]["in"], "kassa_out": data["kassa"]["out"],
            "in": data["total"]["in"], "out": data["total"]["out"],
        })
        ws.cell(r + 1, 2, "ОСТАТОК на конец периода").font = TOTAL_FONT
        cell = ws.cell(r + 1, 7, data["total"]["end"])
        cell.font, cell.number_format = TOTAL_FONT, MONEY

    if "expenses" in which:
        data = await R.expenses_by_code(year=year, month=month, division=None, with_zero=True, _=None, db=db)
        _sheet(wb, "ВСЕГО расходы", f"РАСХОДЫ {title}", EXPENSE_COLS)
        ws = wb["ВСЕГО расходы"]
        r = _fill(ws, EXPENSE_COLS, data["rows"])
        _totals(ws, EXPENSE_COLS, r, data["totals"])

    if "ledger" in which:
        for key, label in LEDGERS:
            data = await R.ledger_report(ledger=key, year=year, month=month, _=None, db=db)
            if not data["rows"]:
                continue
            name = f"Дт Кт {label}"[:31]
            _sheet(wb, name, f"ИНФОРМАЦИЯ о дебиторской и кредиторской задолженности — {label} {title}", LEDGER_COLS)
            ws = wb[name]
            r = _fill(ws, LEDGER_COLS, data["rows"])
            _totals(ws, LEDGER_COLS, r, data["totals"])

    if "daily" in which:
        data = await R.daily_balance(year=year, month=month, currency="UZS", division=None, _=None, db=db)
        c = cols(("date", "Дата", 14), ("income", "Приход", 18, True),
                 ("expense", "Расход", 18, True), ("closing", "Остаток", 20, True))
        c += [{"key": col["key"], "label": col["label"], "width": 20, "money": True}
              for col in data["columns"]]
        _sheet(wb, "ОСТАТОК UZS", f"Остаток денежных средств по дням {title}", c)
        ws = wb["ОСТАТОК UZS"]
        flat = [{"date": "Остаток на начало", "closing": data["opening"]["total"],
                 **data["opening"]["cols"]}]
        flat += [{**r_, **r_["cols"]} for r_ in data["rows"]]
        _fill(ws, c, flat)

    if "fx" in which:
        data = await R.fx_difference(year=year, month=month, _=None, db=db)
        c = cols(("name", "Наименование", 52), ("income", "Доходы", 20, True),
                 ("loss", "Убытки", 20, True))
        _sheet(wb, "Курсовая разница", f"Расчёт валютных курсовых разниц {title}", c)
        ws = wb["Курсовая разница"]
        r = _fill(ws, c, data["rows"])
        _totals(ws, c, r, {"income": data["total_income"], "loss": data["total_loss"]})

    if "taxes" in which:
        data = await R.taxes_report(year=year, month=month, _=None, db=db)
        c = cols(("name", "Наименование налогов", 44),
                 ("debt_start", "Задолженность на начало", 22, True),
                 ("accrued", "Начислено", 20, True), ("paid", "Оплачено", 20, True),
                 ("debt_end", "Задолженность на конец", 22, True),
                 ("overpay", "Переплата", 18, True))
        _sheet(wb, "Налоги", f"Состояние задолженности по видам налогов {title}", c)
        ws = wb["Налоги"]
        r = _fill(ws, c, data["rows"])
        _totals(ws, c, r, data["totals"])

    if "loans" in which:
        data = await R.loans_report(year=year, month=month, _=None, db=db)
        c = cols(("name", "Наименование организации", 42),
                 ("open_debit", "На начало: Дебет", 18, True),
                 ("open_credit", "На начало: Кредит", 18, True),
                 ("turn_debit", "Оборот: Дебет", 18, True),
                 ("turn_credit", "Оборот: Кредит", 18, True),
                 ("end_debit", "На конец: Дебет", 18, True),
                 ("end_credit", "На конец: Кредит", 18, True))
        _sheet(wb, "Займы", f"Задолженность по выданным и полученным займам {title}", c)
        ws = wb["Займы"]
        r = _fill(ws, c, data["rows"])
        _totals(ws, c, r, data["totals"])

    if "gp" in which:
        data = await R.gp_turnover(year=year, month=month, division=None, by_division=True, _=None, db=db)
        c = cols(("division", "Объект", 18), ("code", "Код", 10),
                 ("name", "Наименование товаров", 34), ("unit", "Ед.изм.", 10),
                 ("open_qty", "Остаток нач: Кол-во", 18), ("open_val", "Остаток нач: Сумма", 20, True),
                 ("prod_qty", "Произведено: Кол-во", 18), ("prod_val", "Произведено: Сумма", 20, True),
                 ("sold_qty", "Реализовано: Кол-во", 18), ("sold_val", "Реализовано: Сумма", 20, True),
                 ("close_qty", "Остаток кон: Кол-во", 18), ("close_val", "Остаток кон: Сумма", 20, True))
        _sheet(wb, "ГП оборот", f"Оборот готовой продукции {title}", c)
        _fill(wb["ГП оборот"], c, data["rows"])

    if "materials" in which:
        for kind, label in (("raw", "Склад сырья"), ("spare", "Склад запчастей")):
            data = await R.materials_turnover(kind=kind, year=year, month=month, division=None, by_division=True, _=None, db=db)
            if not data["rows"]:
                continue
            c = cols(("division", "Объект", 18), ("code", "Код", 10),
                     ("name", "Наименование", 36), ("unit", "Ед.изм.", 10),
                     ("open_qty", "Остаток нач: Кол-во", 18), ("open_val", "Остаток нач: Сумма", 20, True),
                     ("recv_qty", "Приход: Кол-во", 18), ("recv_val", "Приход: Сумма", 20, True),
                     ("iss_qty", "Расход: Кол-во", 18), ("iss_val", "Расход: Сумма", 20, True),
                     ("close_qty", "Остаток кон: Кол-во", 18), ("close_val", "Остаток кон: Сумма", 20, True))
            _sheet(wb, label, f"Оборот на складе по дробилкам {title}", c)
            _fill(wb[label], c, data["rows"])

    if "cost" in which:
        data = await R.cost_report(year=year, month=month, division=None, _=None, db=db)
        c = cols(("code", "Код", 10), ("name", "Марка", 28), ("unit", "Ед.изм.", 10),
                 ("produced", "Объём", 14),
                 ("raw_cost", "В С/С: сырьё", 20, True),
                 ("spare_cost", "В С/С: запчасти", 18, True),
                 ("fuel_cost", "В С/С: солярка", 18, True),
                 ("overhead", "В С/С: общие производственные", 22, True),
                 ("unit_cost", "С/С-ТЬ 1 ед.", 18, True),
                 ("sell_unit", "Сверх с/с: реализация", 20, True),
                 ("admin_unit", "Сверх с/с: административные", 22, True),
                 ("other_unit", "Сверх с/с: прочие операционные", 22, True),
                 ("tax_unit", "Сверх с/с: налог на прибыль", 22, True),
                 ("full_unit_cost", "Итого стоимость", 20, True),
                 ("avg_price", "Сред. продажа", 20, True), ("diff", "Разница", 18, True))
        _sheet(wb, "С-сть ГП", f"Себестоимость готовой продукции {title}", c)
        _fill(wb["С-сть ГП"], c, data["rows"])

    if "payroll" in which:
        period = f"{year}-{int(month):02d}" if year and month else None
        data = await payroll_summary(period=period, _=None, db=db)
        c = cols(("division", "Объект", 20), ("expense_code", "Код расхода", 16),
                 ("headcount", "Кол-во", 10), ("gross", "Начислено", 20, True),
                 ("ndfl", "НДФЛ", 18, True), ("inps", "ИНПС", 16, True), ("esp", "ЕСП", 18, True),
                 ("net", "К выдаче", 20, True), ("avans", "Аванс", 18, True),
                 ("paid_cash", "Через кассу", 18, True), ("paid_card", "На карту", 18, True),
                 ("balance", "Долг на конец", 20, True),
                 ("total_cost", "Расходы на сотрудников", 22, True))
        _sheet(wb, "Зарплата свод", f"Свод по заработной плате {title}", c)
        ws = wb["Зарплата свод"]
        r = _fill(ws, c, data["rows"])
        _totals(ws, c, r, data["totals"])


ALL_REPORTS = {"pnl", "balance", "cashflow", "expenses", "ledger", "daily", "fx",
               "taxes", "loans", "gp", "materials", "cost", "payroll"}


@router.get("/reports")
async def export_reports(
    year: int | None = None,
    month: int | None = None,
    only: str | None = None,
    _: User = Depends(require("reports:export")),
    db: AsyncSession = Depends(get_db),
):
    """Книга со всеми отчётами (или только с указанными через `only=pnl,balance`)."""
    which = ALL_REPORTS if not only else {x.strip() for x in only.split(",") if x.strip()}
    unknown = which - ALL_REPORTS
    if unknown:
        raise HTTPException(400, detail=f"Неизвестные отчёты: {', '.join(sorted(unknown))}")
    wb = Workbook()
    await _report_sheets(wb, db, year, month, which)
    if len(wb.sheetnames) == 1 and wb.sheetnames[0] == "Sheet":
        raise HTTPException(404, detail="За выбранный период нет данных")
    suffix = f"{year}-{int(month):02d}" if year and month else (str(year) if year else "все")
    return _xlsx(wb, f"PROFIT_DIVIDER_отчёты_{suffix}.xlsx")


# ============================ РЕЕСТРЫ ============================
@router.get("/registry/{name}")
async def export_registry(
    name: str,
    year: int | None = None,
    month: int | None = None,
    current: User = Depends(require("reports:export")),
    db: AsyncSession = Depends(get_db),
):
    """Выгрузка первичного реестра: transactions | receipts | issues |
    productions | sales | services | payroll | employees."""
    title = _period_title(year, month)
    wb = Workbook()

    def money(v) -> float:
        return round(float(v or 0), 2)

    if name == "transactions":
        stmt = select(Transaction).options(selectinload(Transaction.organization)).order_by(
            Transaction.doc_date, Transaction.id
        )
        rows_db = (await db.execute(R._period(stmt, year, month))).scalars().all()
        c = cols(("doc_date", "Дата", 13), ("account", "Счёт", 10), ("direction", "Тип", 12),
                 ("org", "Отправитель/Получатель", 40), ("corr_inn", "ИНН", 14),
                 ("doc_no", "№ документа", 14), ("mfo", "МФО", 10),
                 ("corr_account", "Счёт корресп.", 24), ("corr_name", "Наименование корресп.", 32),
                 ("purpose", "Назначение по кодировке", 28),
                 ("expense_code", "Код расхода", 14), ("cashflow_code", "Код cash flow", 14),
                 ("division", "Объект", 14), ("currency", "Валюта", 10),
                 ("amount", "Сумма", 18, True), ("rate", "Курс", 12),
                 ("amount_uzs", "Сумма UZS", 20, True), ("amount_usd", "Сумма USD", 20, True),
                 ("description", "Наименование платежа", 46))
        _sheet(wb, "БАНК-КАССА", f"Реестр операций {title}", c)
        _fill(wb["БАНК-КАССА"], c, [{
            "doc_date": t.doc_date, "account": "БАНК" if t.account == "bank" else "КАССА",
            "direction": "Приход" if t.direction == "income" else "Расход",
            "org": t.organization.name if t.organization else "",
            "corr_inn": t.corr_inn, "doc_no": t.doc_no, "mfo": t.mfo,
            "corr_account": t.corr_account, "corr_name": t.corr_name, "purpose": t.purpose,
            "expense_code": t.expense_code, "cashflow_code": t.cashflow_code,
            "division": t.division, "currency": t.currency, "amount": money(t.amount),
            "rate": float(t.rate or 1), "amount_uzs": money(t.amount_uzs),
            "amount_usd": money(t.amount_usd), "description": t.description,
        } for t in rows_db])

    elif name == "receipts":
        stmt = select(MaterialReceipt).options(
            selectinload(MaterialReceipt.material), selectinload(MaterialReceipt.organization)
        ).order_by(MaterialReceipt.doc_date, MaterialReceipt.id)
        rows_db = (await db.execute(R._period(stmt, year, month, MaterialReceipt.doc_date))).scalars().all()
        c = cols(("doc_date", "Дата", 13), ("org", "Наименование поставщика", 40),
                 ("code", "Код сырья", 14), ("material", "Наименование материала", 36),
                 ("unit", "Ед. изм.", 10), ("division", "Дробилка", 14),
                 ("payment_type", "Вид оплаты", 18), ("vat", "Плательщик НДС", 16),
                 ("qty", "Кол-во", 14), ("price_uzs", "Цена (без учета НДС)", 20, True),
                 ("amount_uzs", "Сумма без учета НДС", 22, True),
                 ("vat_amount", "Сумма НДС", 18, True),
                 ("amount_gross", "Сумма с учетом НДС", 22, True),
                 ("note", "Примечание", 30))
        _sheet(wb, "Приход сырья", f"Оприходование сырья и запчастей {title}", c)
        _fill(wb["Приход сырья"], c, [{
            "doc_date": x.doc_date, "org": x.organization.name if x.organization else "",
            "code": x.material.code if x.material else "",
            "material": x.material.name if x.material else "",
            "unit": x.material.unit if x.material else "", "division": x.division,
            "payment_type": x.payment_type or "",
            "qty": float(x.qty or 0), "price_uzs": money(x.price_uzs),
            "amount_uzs": money(x.amount_uzs), "vat_amount": money(x.vat_amount),
            "amount_gross": money(x.amount_gross),
            "vat": "с учетом НДС" if x.vat else "без учета НДС", "note": x.note,
        } for x in rows_db])

    elif name == "issues":
        stmt = select(MaterialIssue).options(selectinload(MaterialIssue.material)).order_by(
            MaterialIssue.doc_date, MaterialIssue.id
        )
        rows_db = (await db.execute(R._period(stmt, year, month, MaterialIssue.doc_date))).scalars().all()
        c = cols(("doc_date", "Дата", 13), ("code", "Код сырья", 14),
                 ("material", "Наименование", 36), ("unit", "Ед. изм.", 10),
                 ("division", "Объект", 14), ("expense_code", "Код расхода", 14),
                 ("qty", "Кол-во", 14), ("price_uzs", "Цена UZS", 18, True),
                 ("cost_uzs", "Сумма UZS", 20, True),
                 ("note", "Примечание", 30))
        _sheet(wb, "Расход сырья", f"Расход сырья и запчастей {title}", c)
        _fill(wb["Расход сырья"], c, [{
            "doc_date": x.doc_date, "code": x.material.code if x.material else "",
            "material": x.material.name if x.material else "",
            "unit": x.material.unit if x.material else "",
            "division": x.division, "expense_code": x.expense_code,
            "qty": float(x.qty or 0),
            "price_uzs": money(float(x.cost_uzs or 0) / float(x.qty)) if float(x.qty or 0) else 0,
            "cost_uzs": money(x.cost_uzs), "note": x.note,
        } for x in rows_db])

    elif name == "productions":
        stmt = select(Production).options(selectinload(Production.product)).order_by(
            Production.doc_date, Production.id
        )
        rows_db = (await db.execute(R._period(stmt, year, month, Production.doc_date))).scalars().all()
        c = cols(("doc_date", "Дата", 13), ("division", "Объект", 16), ("code", "Код товара", 14),
                 ("product", "Наименование ГП", 34), ("unit", "Ед. изм.", 10),
                 ("qty", "Кол-во", 14), ("unit_cost", "Цена (с-сть)", 18, True),
                 ("amount_uzs", "Сумма (с-сть)", 20, True), ("note", "Примечание", 30))
        _sheet(wb, "Производство ГП", f"Произведённая готовая продукция {title}", c)
        _fill(wb["Производство ГП"], c, [{
            "doc_date": x.doc_date, "division": x.division,
            "code": x.product.code if x.product else "",
            "product": x.product.name if x.product else "",
            "unit": x.product.unit if x.product else "",
            "qty": float(x.qty or 0), "unit_cost": money(x.unit_cost),
            "amount_uzs": money(x.amount_uzs), "note": x.note,
        } for x in rows_db])

    elif name == "sales":
        stmt = select(Sale).options(
            selectinload(Sale.product), selectinload(Sale.organization)
        ).order_by(Sale.doc_date, Sale.id)
        rows_db = (await db.execute(R._period(stmt, year, month, Sale.doc_date))).scalars().all()
        c = cols(("doc_date", "Дата", 13), ("inn", "ИНН", 14),
                 ("org", "Наименование клиента", 40), ("division", "Дробилка", 14),
                 ("code", "Код товара", 12), ("product", "Наименование ГП", 30),
                 ("payment_type", "Вид оплаты", 18), ("unit", "Ед.изм.", 10),
                 ("qty", "Кол-во", 14), ("price_uzs", "Цена с НДС", 18, True),
                 ("gross", "Сумма с учетом НДС", 22, True), ("vat_amount", "Сумма НДС", 18, True),
                 ("revenue_net", "Сумма без учета НДС", 22, True),
                 ("cogs_uzs", "Себестоимость", 20, True))
        _sheet(wb, "Продажа ГП", f"Реализация готовой продукции {title}", c)
        _fill(wb["Продажа ГП"], c, [{
            "doc_date": x.doc_date, "inn": x.organization.inn if x.organization else "",
            "org": x.organization.name if x.organization else "",
            "division": x.division, "code": x.product.code if x.product else "",
            "product": x.product.name if x.product else "",
            "payment_type": x.payment_type or "",
            "unit": x.product.unit if x.product else "", "qty": float(x.qty or 0),
            "price_uzs": money(x.price_uzs),
            "gross": money(float(x.qty or 0) * float(x.price_uzs or 0)),
            "vat_amount": money(x.vat_amount), "revenue_net": money(x.revenue_net),
            "cogs_uzs": money(x.cogs_uzs),
        } for x in rows_db])

    elif name == "services":
        stmt = select(Service).options(selectinload(Service.organization)).order_by(
            Service.doc_date, Service.id
        )
        rows_db = (await db.execute(R._period(stmt, year, month, Service.doc_date))).scalars().all()
        for direction, label in (("received", "Полученные УСЛУГИ"), ("provided", "Оказанные УСЛУГИ")):
            subset = [x for x in rows_db if x.direction == direction]
            if not subset:
                continue
            c = cols(("doc_date", "Дата счёт-фактуры", 18), ("inn", "ИНН", 14),
                     ("org", "Наименование организации", 40), ("service_type", "Вид услуг", 30),
                     ("expense_code", "Код платежа", 14), ("division", "Объект", 14),
                     ("net", "Сумма без НДС", 20, True), ("vat_amount", "Сумма НДС", 18, True),
                     ("amount", "Сумма с НДС", 20, True), ("note", "Назначение", 34))
            _sheet(wb, label, f"Информация об услугах {title}", c)
            _fill(wb[label], c, [{
                "doc_date": x.doc_date, "inn": x.organization.inn if x.organization else "",
                "org": x.organization.name if x.organization else "",
                "service_type": x.service_type, "expense_code": x.expense_code,
                "division": x.division, "net": money(x.net), "vat_amount": money(x.vat_amount),
                "amount": money(x.amount), "note": x.note,
            } for x in subset])

    elif name == "payroll":
        stmt = select(PayrollEntry).options(selectinload(PayrollEntry.employee)).order_by(
            PayrollEntry.id
        )
        if year and month:
            stmt = stmt.where(PayrollEntry.period == f"{year}-{int(month):02d}")
        rows_db = (await db.execute(stmt)).scalars().all()
        c = cols(("full_name", "Ф.И.О.", 34), ("division", "Объект", 14),
                 ("department", "Отдел", 12), ("position", "Должность", 28),
                 ("expense_code", "Код расхода", 14),
                 ("norm_days", "Норма дней", 12), ("worked_days", "Отработано", 12),
                 ("overtime_days", "Сверхурочные", 14),
                 ("debt_start", "Долг на начало", 18, True),
                 ("oklad", "Оклад", 18, True), ("nadbavka", "Надбавка", 16, True),
                 ("pitanie", "Питание", 16, True), ("bonus", "Премия", 16, True),
                 ("benzin", "Бензин", 16, True), ("other_accrued", "Прочие начисления", 18, True),
                 ("gross", "Всего начислено", 20, True),
                 ("ndfl", "НДФЛ", 16, True), ("inps", "ИНПС", 14, True),
                 ("hold_pitanie", "Удерж. питание", 16, True),
                 ("hold_alimony", "Алименты", 16, True), ("hold_other", "Прочие удерж.", 16, True),
                 ("fine", "Штраф", 14, True), ("withheld", "Всего удержано", 20, True),
                 ("net", "Сумма к выдаче", 20, True), ("avans", "Аванс", 16, True),
                 ("paid_cash", "Через кассу", 16, True), ("paid_card", "На карту", 16, True),
                 ("paid", "Выплачено всего", 20, True), ("balance", "Долг на конец", 18, True),
                 ("esp", "ЕСП", 16, True), ("total_cost", "Расходы на сотрудника", 22, True))
        _sheet(wb, "Зарплата", f"Расчёт заработной платы {title}", c)
        _fill(wb["Зарплата"], c, [{
            "full_name": x.employee.full_name if x.employee else "",
            "division": x.employee.division if x.employee else "",
            "department": x.employee.department if x.employee else "",
            "position": x.employee.position if x.employee else "",
            "expense_code": x.employee.expense_code if x.employee else "",
            **{k: money(getattr(x, k)) for k in (
                "debt_start", "oklad", "nadbavka", "pitanie", "bonus", "benzin",
                "other_accrued", "gross", "ndfl", "inps", "hold_pitanie", "hold_alimony",
                "hold_other", "fine", "withheld", "net", "avans", "paid_cash", "paid_card",
                "paid", "balance", "esp", "total_cost")},
            "norm_days": float(x.norm_days or 0), "worked_days": float(x.worked_days or 0),
            "overtime_days": float(x.overtime_days or 0),
        } for x in rows_db])

    elif name == "employees":
        rows_db = (await db.execute(select(Employee).order_by(Employee.full_name))).scalars().all()
        c = cols(("full_name", "Ф.И.О.", 34), ("inn", "ИНН сотрудника", 16),
                 ("division", "Объект", 14), ("department", "Отдел", 12),
                 ("position", "Должность", 30), ("category", "Категория", 24),
                 ("group", "Группа", 24), ("status", "Статус", 16), ("state", "Состояние", 24),
                 ("expense_code", "Код расхода", 14), ("payment_type", "Вид выплаты", 14),
                 ("salary", "Оклад", 20, True), ("is_active", "Активен", 12))
        _sheet(wb, "Сотрудники", "Список сотрудников", c)
        _fill(wb["Сотрудники"], c, [{
            "full_name": e.full_name, "inn": e.inn, "division": e.division,
            "department": e.department, "position": e.position, "category": e.category,
            "group": e.group, "status": e.status, "state": e.state,
            "expense_code": e.expense_code, "payment_type": e.payment_type,
            "salary": money(e.salary), "is_active": "да" if e.is_active else "нет",
        } for e in rows_db])

    else:
        raise HTTPException(404, detail=f"Неизвестный реестр: {name}")

    suffix = f"_{year}-{int(month):02d}" if year and month else ""
    return _xlsx(wb, f"PROFIT_DIVIDER_{name}{suffix}.xlsx")
