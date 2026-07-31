"""Загрузка первичных данных из Excel (.xlsx).

Файл читается по ЗАГОЛОВКАМ первой строки — порядок колонок не важен,
лишние колонки игнорируются. Поддерживаются как русские названия из книги
PROFIT DIVIDER, так и англоязычные ключи (для файлов, выгруженных из системы).

Всегда доступен предпросмотр: `dry_run=true` разбирает файл и возвращает,
что будет создано и какие строки содержат ошибки — ничего не записывая.
"""
from datetime import date, datetime
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ..database import get_db
from ..events import record
from ..ledger import recompute_org_balances
from ..models import (
    BankAccount,
    CashRegister,
    Material,
    MaterialReceipt,
    Organization,
    Product,
    Sale,
    Transaction,
    User,
)
from ..rates import get_rates
from ..security import require
from .inventory import recompute_material, recompute_product
from .transactions import _resolve_rate, _usd, _uzs

router = APIRouter(prefix="/api/import", tags=["import"])

MAX_ROWS = 20000


def _norm(v) -> str:
    return " ".join(str(v).strip().lower().split()) if v is not None else ""


def _read(file_bytes: bytes) -> list[dict]:
    """Прочитать первый лист в список словарей по заголовкам первой строки."""
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(400, detail=f"Не удалось прочитать файл: {exc}") from None
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        raise HTTPException(400, detail="Файл пуст") from None
    keys = [_norm(h) for h in header]
    if not any(keys):
        raise HTTPException(400, detail="В первой строке нет заголовков колонок")
    rows = []
    for i, raw in enumerate(it, 2):
        if not any(v is not None and str(v).strip() != "" for v in raw):
            continue
        rows.append({"_row": i, **{k: v for k, v in zip(keys, raw) if k}})
        if len(rows) > MAX_ROWS:
            raise HTTPException(400, detail=f"Слишком много строк (максимум {MAX_ROWS})")
    return rows


def _pick(row: dict, *names):
    for n in names:
        v = row.get(_norm(n))
        if v is not None and str(v).strip() != "":
            return v
    return None


def _num(row: dict, *names, default=0.0) -> float:
    v = _pick(row, *names)
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    txt = str(v).replace(" ", "").replace(" ", "").replace(",", ".")
    try:
        return float(txt)
    except ValueError:
        raise ValueError(f"«{v}» — не число") from None


def _date(row: dict, *names) -> date:
    v = _pick(row, *names)
    if v is None:
        raise ValueError("не указана дата")
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    txt = str(v).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(txt, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"«{v}» — непонятная дата (ожидается ГГГГ-ММ-ДД или ДД.ММ.ГГГГ)")


def _flag(row: dict, *names) -> bool:
    v = _pick(row, *names)
    txt = _norm(v)
    return txt in ("да", "yes", "true", "1", "с учетом ндс", "с ндс", "+")


class Resolver:
    """Поиск справочных записей по названию/ИНН/коду с кэшем."""

    def __init__(self, orgs, materials, products, banks, tills):
        self.orgs_name = {(o.name or "").upper(): o for o in orgs}
        self.orgs_inn = {(o.inn or ""): o for o in orgs if o.inn}
        self.mat_name = {(m.name or "").upper(): m for m in materials}
        self.mat_code = {(m.code or ""): m for m in materials if m.code}
        self.prod_name = {(p.name or "").upper(): p for p in products}
        self.prod_code = {(p.code or ""): p for p in products if p.code}
        self.banks = {(b.name or "").upper(): b for b in banks}
        self.tills = {(c.name or "").upper(): c for c in tills}

    def org(self, name, inn):
        if inn and str(inn).strip() in self.orgs_inn:
            return self.orgs_inn[str(inn).strip()]
        if name and str(name).strip().upper() in self.orgs_name:
            return self.orgs_name[str(name).strip().upper()]
        return None

    def material(self, name, code):
        if code and str(code).strip() in self.mat_code:
            return self.mat_code[str(code).strip()]
        if name and str(name).strip().upper() in self.mat_name:
            return self.mat_name[str(name).strip().upper()]
        return None

    def product(self, name, code):
        if code and str(code).strip() in self.prod_code:
            return self.prod_code[str(code).strip()]
        if name and str(name).strip().upper() in self.prod_name:
            return self.prod_name[str(name).strip().upper()]
        return None


async def _resolver(db: AsyncSession) -> Resolver:
    async def all_of(model):
        return (await db.execute(select(model))).scalars().all()

    return Resolver(
        await all_of(Organization), await all_of(Material), await all_of(Product),
        await all_of(BankAccount), await all_of(CashRegister),
    )


TEMPLATES = {
    "transactions": {
        "label": "БАНК / КАССА — операции",
        "columns": [
            "Дата*", "Счёт (БАНК/КАССА)*", "Тип (Приход/Расход)*", "Валюта", "Сумма*",
            "ИНН", "Отправитель/Получатель", "Код расхода", "Код cash flow", "Объект",
            "№ документа", "МФО", "Счёт корресп.", "Наименование корресп.",
            "Назначение по кодировке", "Наименование платежа", "Банковский счёт", "Касса",
        ],
    },
    "receipts": {
        "label": "Приход сырья и запчастей",
        "columns": ["Дата*", "ИНН", "Наименование поставщика", "Код сырья",
                    "Наименование материала*", "Дробилка", "Кол-во*",
                    "Цена (без НДС)*", "НДС", "Примечание"],
    },
    "sales": {
        "label": "Продажа готовой продукции",
        "columns": ["Дата*", "ИНН", "Покупатель", "Код товара", "Наименование ГП*",
                    "Объект", "Кол-во*", "Цена с НДС*", "НДС", "Примечание"],
    },
}


@router.get("/templates")
async def templates(_: User = Depends(require("transactions:create"))):
    """Какие колонки ждёт загрузчик для каждого типа файла."""
    return [{"key": k, **v} for k, v in TEMPLATES.items()]


@router.post("/{kind}")
async def import_file(
    kind: str,
    file: UploadFile = File(...),
    dry_run: bool = Query(True, description="только проверить файл, ничего не записывать"),
    current: User = Depends(require("transactions:create")),
    db: AsyncSession = Depends(get_db),
):
    if kind not in TEMPLATES:
        raise HTTPException(404, detail=f"Неизвестный тип загрузки: {kind}")
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, detail="Ожидается файл .xlsx")

    rows = _read(await file.read())
    res = await _resolver(db)
    nds = (await get_rates(db))["nds_rate"]

    created, errors = [], []
    touched_orgs: set[int] = set()
    touched_mats: set[int] = set()
    touched_prods: set[int] = set()

    for row in rows:
        n = row["_row"]
        try:
            if kind == "transactions":
                obj = await _tx_row(db, row, res)
            elif kind == "receipts":
                obj = _receipt_row(row, res)
            elif kind == "sales":
                obj = _sale_row(row, res)
            created.append({"row": n, "summary": obj["summary"]})
            if not dry_run:
                db.add(obj["model"])
                for bucket, key in (
                    (touched_orgs, "organization_id"),
                    (touched_mats, "material_id"),
                    (touched_prods, "product_id"),
                ):
                    v = getattr(obj["model"], key, None)
                    if v:
                        bucket.add(v)
        except (ValueError, LookupError) as exc:
            errors.append({"row": n, "error": str(exc)})

    if dry_run or errors:
        return {
            "dry_run": True, "applied": False,
            "total": len(rows), "ok": len(created), "failed": len(errors),
            "errors": errors[:200], "preview": created[:50],
            "message": (
                "Найдены ошибки — ничего не загружено. Исправьте строки и повторите."
                if errors else "Проверка пройдена. Повторите с dry_run=false для загрузки."
            ),
        }

    await db.flush()
    for mid in touched_mats:
        await recompute_material(db, mid)
    for pid in touched_prods:
        await recompute_product(db, pid)
    if touched_orgs:
        await recompute_org_balances(db, list(touched_orgs))
    await db.commit()
    await record(db, current, "create", f"import:{kind}", f"{len(created)} строк")
    return {
        "dry_run": False, "applied": True,
        "total": len(rows), "ok": len(created), "failed": 0,
        "errors": [], "preview": created[:50],
        "message": f"Загружено строк: {len(created)}",
    }


async def _tx_row(db: AsyncSession, row: dict, res: Resolver) -> dict:
    d = _date(row, "Дата", "doc_date")
    acc = _norm(_pick(row, "Счёт (БАНК/КАССА)", "Счёт", "account"))
    account = "kassa" if acc.startswith(("касса", "kassa", "kassa")) else "bank"
    typ = _norm(_pick(row, "Тип (Приход/Расход)", "Тип", "direction"))
    if typ.startswith(("приход", "income", "kirim")):
        direction = "income"
    elif typ.startswith(("расход", "expense", "chiqim")):
        direction = "expense"
    else:
        raise ValueError("тип операции: «Приход» или «Расход»")
    currency = (str(_pick(row, "Валюта", "currency") or "UZS")).strip().upper()
    if currency not in ("UZS", "USD"):
        raise ValueError(f"валюта «{currency}» не поддерживается (UZS или USD)")
    amount = _num(row, "Сумма", "amount")
    if amount <= 0:
        raise ValueError("сумма должна быть больше нуля")
    rate = await _resolve_rate(db, currency, d)

    org = res.org(_pick(row, "Отправитель/Получатель", "Организация", "org"),
                  _pick(row, "ИНН", "ИНН по данным банка", "inn"))
    bank = res.banks.get(str(_pick(row, "Банковский счёт", "bank_account") or "").upper())
    till = res.tills.get(str(_pick(row, "Касса", "cash_register") or "").upper())

    # Заголовки в книге PROFIT DIVIDER на листах БАНК и КАССА названы по-разному
    # («код расходов» / «Код для расходов», «код cash flow» / «Код платежа»),
    # поэтому у каждого поля перечислены оба варианта — иначе колонка молча
    # не подхватывалась и код расхода терялся при загрузке.
    tx = Transaction(
        doc_date=d, direction=direction, account=account, currency=currency,
        amount=amount, rate=rate,
        amount_usd=_usd(currency, amount, rate), amount_uzs=_uzs(currency, amount, rate),
        organization_id=org.id if org else None,
        bank_account_id=bank.id if bank else None,
        cash_register_id=till.id if till else None,
        expense_code=str(_pick(row, "Код расхода", "код расходов",
                              "Код для расходов", "expense_code") or ""),
        cashflow_code=str(_pick(row, "Код cash flow", "код cash flow",
                                "Код платежа", "cashflow_code") or ""),
        division=str(_pick(row, "Объект", "Подразделение", "division") or ""),
        doc_no=str(_pick(row, "№ документа", "Номер документа", "doc_no") or ""),
        mfo=str(_pick(row, "МФО", "МФО корресп.", "mfo") or ""),
        corr_account=str(_pick(row, "Счёт корресп.", "Счет корреспондента",
                               "corr_account") or ""),
        corr_name=str(_pick(row, "Наименование корресп.", "corr_name") or ""),
        corr_inn=str(_pick(row, "ИНН по данным банка", "ИНН", "corr_inn") or ""),
        purpose=str(_pick(row, "Назначение по кодировке", "Назначение", "purpose") or ""),
        description=str(_pick(row, "Наименование платежа", "description") or ""),
    )
    return {"model": tx, "summary": f"{d} {account} {direction} {amount:,.2f} {currency}"}


def _receipt_row(row: dict, res: Resolver) -> dict:
    d = _date(row, "Дата", "doc_date")
    name = _pick(row, "Наименование материала", "Наименование сырья", "material")
    mat = res.material(name, _pick(row, "Код сырья", "code"))
    if not mat:
        raise LookupError(f"материал «{name}» не найден в справочнике сырья/запчастей")
    qty = _num(row, "Кол-во", "qty")
    if qty <= 0:
        raise ValueError("количество должно быть больше нуля")
    price = _num(row, "Цена (без НДС)", "Цена", "price_uzs")
    org = res.org(_pick(row, "Наименование поставщика", "Поставщик", "org"),
                  _pick(row, "ИНН", "inn"))
    rec = MaterialReceipt(
        doc_date=d, material_id=mat.id, organization_id=org.id if org else None,
        division=str(_pick(row, "Дробилка", "Объект", "division") or ""),
        qty=qty, price_uzs=price, amount_uzs=round(qty * price, 2),
        vat=_flag(row, "НДС", "vat"), note=str(_pick(row, "Примечание", "note") or ""),
    )
    return {"model": rec, "summary": f"{d} {mat.name} {qty} × {price:,.2f}"}


def _sale_row(row: dict, res: Resolver) -> dict:
    d = _date(row, "Дата", "doc_date")
    name = _pick(row, "Наименование ГП", "Наименование продукции", "product")
    prod = res.product(name, _pick(row, "Код товара", "code"))
    if not prod:
        raise LookupError(f"продукция «{name}» не найдена в справочнике ГП")
    qty = _num(row, "Кол-во", "qty")
    if qty <= 0:
        raise ValueError("количество должно быть больше нуля")
    price = _num(row, "Цена с НДС", "Цена", "price_uzs")
    org = res.org(_pick(row, "Покупатель", "Организация", "org"), _pick(row, "ИНН", "inn"))
    sale = Sale(
        doc_date=d, product_id=prod.id, organization_id=org.id if org else None,
        division=str(_pick(row, "Объект", "division") or ""),
        qty=qty, price_uzs=price, vat=_flag(row, "НДС", "vat"),
        note=str(_pick(row, "Примечание", "note") or ""),
    )
    return {"model": sale, "summary": f"{d} {prod.name} {qty} × {price:,.2f}"}
