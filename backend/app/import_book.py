"""Разовая загрузка книги «Баланс … PROFIT-DIVIDER.xlsx» в базу.

Запуск внутри контейнера:

    docker exec finasist_backend python -m app.import_book /tmp/book.xlsx \
        [--wipe] [--rates /tmp/rates.json]

`--wipe` сначала удаляет ВСЕ первичные документы (операции, приходы, расходы,
производство, продажи, услуги, зарплату, займы, налоги, закрытые месяцы) —
справочники (организации, сырьё, ГП, коды, пользователи) остаются.

`--rates` — JSON вида {"2026-07-01": 12009.27, ...}. В самой книге на листе
«Курс доллара» на весь июль проставлена единица, то есть валютная колонка
повторяет сумовую. С этим файлом курс берётся из него (мы кладём туда курс ЦБ
РУз), и валютные суммы считаются по-настоящему. Сумовые суммы и баланс от
курса не зависят — меняются только колонки «в долларах» и курсовая разница.

Что читается из книги:
  · «Курс доллара»        -> exchange_rates, если не передан --rates
  · INFO + «ОСТАТОК UZS»  -> bank_accounts с входящими остатками
  · «КАССА» (колонка A)   -> cash_registers, входящий остаток общий
  · «Дт Кт …» (5 листов)  -> входящее сальдо контрагентов по ведомостям
  · «Займы»               -> loans
  · «БАНК» / «КАССА»      -> transactions
  · «Остаток ГП» / «Остаток сырья и запчастей» -> входящие остатки складов
  · «Производства Приход ГП» -> productions
  · «Налоги»              -> taxes
  · «ОС», «Офис Note»     -> settings (fa_cost, capital_charter)

ВАЖНО про диапазоны: книга местами ссылается на БАНК!$9:$350 и $9:$359, хотя
данные идут до 704-й строки. Загрузчик читает ВСЕ строки с датой, поэтому в
системе оборот полный — из-за этого итоги могут не совпасть с ячейками книги,
которые считались по обрезанному диапазону.
"""
import asyncio
import calendar
import json
import sys
from datetime import date, datetime

from openpyxl import load_workbook
from sqlalchemy import delete, select, text

from .database import AsyncSessionLocal
from .ledger import recompute_org_balances
from .models import (
    BankAccount,
    CashRegister,
    Employee,
    ExchangeRate,
    Loan,
    LoanEntry,
    Material,
    MaterialIssue,
    MaterialReceipt,
    MaterialStock,
    Organization,
    PayrollEntry,
    PeriodClose,
    PeriodSetting,
    Product,
    ProductPrice,
    ProductStock,
    Production,
    Sale,
    Service,
    Setting,
    Tax,
    Transaction,
)
from .routers.inventory import recompute_material, recompute_product

BOOK_MONTH = 7
BOOK_YEAR = 2026
OPENING_DATE = date(2026, 6, 30)   # входящие сальдо: на конец предыдущего месяца
TAX_DATE = date(2026, 7, 31)
# «Производство» и «Расход сырья» в книге месячные, без настоящей даты.
# Расход сырья относим на ПОСЛЕДНИЙ день месяца (а не на 30-е: в 31-дневном
# месяце это ещё не конец, и приходы 31-го числа не успевали лечь на склад
# ДО расхода — средняя цена уходила в минус и оставалась испорченной до
# конца месяца).
MONTH_END = date(BOOK_YEAR, BOOK_MONTH, calendar.monthrange(BOOK_YEAR, BOOK_MONTH)[1])
# А производство ГП — НАОБОРОТ, на ПЕРВЫЙ день месяца: продажи (реализация)
# идут с настоящими датами весь месяц, и если выпуск встанет 31-го числа,
# все продажи 1-30 числа спишутся с трёх сотен единиц входящего остатка,
# склад уйдёт в глубокий минус, и та же порча средней цены — только теперь
# в 10 раз (обнаружено: средняя цена Клинес на Махстоне вместо ~106 тыс.
# получалась ~1 млн сум). Выпуск должен лечь ДО первой же продажи месяца.
MONTH_START = date(BOOK_YEAR, BOOK_MONTH, 1)

# лист «Дт Кт …» -> ведомость в системе
LEDGER_SHEETS = {
    "Дт Кт поставщ подрядчик": ("suppliers", "supplier"),
    "Дт Кт покуп заказчик": ("customers", "customer"),
    "Дт Кт СЕЙФ (ЯККАСАРОЙ)": ("safe", "other"),
    "Дт Кт прочие (услуги)": ("services", "other"),
    "Дт Кт З.п": ("salary", "other"),
}
SKIP_NAMES = {"всего", "в том числе", "всего по", "всего по дробилки", "итого"}

# «Азмур» в «Приход запчастей» — это не отдельный объект, а место приёмки:
# запчасть приходуется на «Азмур», а расходуется (лист «Расход сырья и
# запчастей») уже на «Махстон». Поэтому в книге одна и та же деталь на конец
# месяца висит излишком на «Азмуре» и точно таким же минусом на «Махстоне» —
# проверено на всех 71 позициях, совпадает без остатка. Приравниваем при
# чтении, чтобы приход и расход считались одним и тем же складом.
DIVISION_ALIASES = {"Азмур": "Махстон"}

# Точечные расхождения книги: «Объект»/касса в строке — один, а суффикс кода
# расходов — другой (книжная «РАСХОДЫ <объект>» считает по коду, не по
# «Объекту»). Оба случая найдены сверкой себестоимости с книгой день в день:
#   КАССА стр.800 — регистратор «Турк», код «2032_М» (расход «Курс разница»);
#   Полученные УСЛУГИ стр.61 — «Объект»=Турк, код «2035_М» (KS PROFIT SOLAR).
# Не общее правило (регистратор/«Объект» в остальных ~660 строках верны) —
# точечная поправка только этих двух строк, определяемых номером строки листа.
ROW_DIVISION_OVERRIDE = {
    ("КАССА", 800): "Махстон",
    ("Полученные УСЛУГИ", 61): "Махстон",
}

def num(v) -> float:
    return float(v) if isinstance(v, (int, float)) else 0.0


def txt(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return " ".join(str(v).split())


def as_date(v):
    if isinstance(v, datetime):
        return v.date()
    return v if isinstance(v, date) else None


def okey(v: str) -> str:
    """Ключ контрагента. В книге это «код по ИНН» — иногда не число, а текст
    («ШАХЗОД СОТУВЧИ под-отчёт»), а колонка `inn` в базе — 20 символов.
    Обрезаем в ОДНОМ месте, чтобы ведомости, реестр и операции сходились."""
    return (v or "").strip()[:20]


def is_summary_row(name: str) -> bool:
    """Строка-итог («ВСЕГО», «в том числе»), а не контрагент."""
    n = name.strip().lower()
    return n in SKIP_NAMES or n.startswith("всего")


def is_total_row(name: str) -> bool:
    return not name.strip() or is_summary_row(name)


class Book:
    """Разобранная книга: всё читается один раз, потом пишется в базу."""

    def __init__(self, path: str):
        self.wb = load_workbook(path, data_only=True, read_only=True)

    def rows(self, sheet: str, first: int, last: int | None = None, width: int = 40):
        ws = self.wb[sheet]
        for i, r in enumerate(ws.iter_rows(min_row=first, max_row=last, max_col=width,
                                           values_only=True), first):
            yield i, r

    # ---------- справочные блоки ----------

    def rates(self) -> dict[date, float]:
        out = {}
        for _i, r in self.rows("Курс доллара", 4, 40, 2):
            d, v = as_date(r[0]), num(r[1])
            if d and v:
                out[d] = v
        return out

    def banks(self) -> list[dict]:
        """INFO A31:E40 — счета; «ОСТАТОК UZS» H4:N4 — входящие остатки."""
        info = {}
        for _i, r in self.rows("INFO", 31, 40, 5):
            name = txt(r[1])
            if name:
                info[name] = {"account_no": txt(r[2]), "mfo": txt(r[3])}
        head = next(r for i, r in self.rows("ОСТАТОК UZS", 3, 3, 20))
        open_row = next(r for i, r in self.rows("ОСТАТОК UZS", 4, 4, 20))
        out = []
        for c in range(7, 14):  # H..N
            name = txt(head[c])
            if not name or name == "0":
                continue
            meta = info.get(name, {})
            out.append({
                "name": name,
                "account_no": meta.get("account_no", ""),
                "mfo": meta.get("mfo", ""),
                "opening": round(num(open_row[c]), 2),
            })
        return out

    def cash_opening(self) -> float:
        """«ОСТАТОК UZS» E4 — общий остаток по кассам на начало."""
        row = next(r for i, r in self.rows("ОСТАТОК UZS", 4, 4, 10))
        return round(num(row[4]), 2)

    def ledger_openings(self) -> list[dict]:
        """Входящее сальдо контрагентов из пяти листов «Дт Кт …».

        Ключ строки (колонка B — «код по ИНН») — это то же самое, по чему листы
        БАНК и КАССА связываются с ведомостью через SUMIFS. Один ключ может
        встретиться в листе дважды (две строки на одного контрагента) — такие
        строки СУММИРУЮТСЯ, иначе вторая затирает первую и сальдо теряется.
        """
        merged: dict[tuple[str, str], dict] = {}
        for sheet, (ledger, category) in LEDGER_SHEETS.items():
            for _i, r in self.rows(sheet, 9, None, 20):
                key, name = txt(r[1]), txt(r[2])
                if is_summary_row(name):
                    continue
                debit, credit = num(r[3]), num(r[5])
                if not key and not name:
                    # в книге встречается строка без ИНН и без названия, но с
                    # суммой (стр. 66 листа поставщиков). Пропустить её — значит
                    # потерять сальдо, поэтому заводим отдельного контрагента
                    if not debit and not credit:
                        continue
                    key = f"{ledger[:4]}-стр{_i}"          # ИНН — не длиннее 20 символов
                    name = f"Без наименования ({sheet}, стр. {_i})"
                key = key or name
                row = merged.setdefault((ledger, key.upper()), {
                    "key": key, "name": name, "ledger": ledger,
                    "category": category, "opening": 0.0,
                })
                row["opening"] = round(row["opening"] + debit - credit, 2)
        return list(merged.values())

    def loans(self) -> list[dict]:
        out = []
        for _i, r in self.rows("Займы", 9, None, 20):
            key, name = txt(r[1]), txt(r[2])
            if is_total_row(name):
                continue
            debit, credit = num(r[3]), num(r[5])
            if debit:
                out.append({"name": name or key, "direction": "given", "opening": debit})
            if credit:
                out.append({"name": name or key, "direction": "received", "opening": credit})
        return out

    def bank_tx(self) -> list[dict]:
        """Лист «БАНК», строки 9.. — все, у которых есть дата."""
        out = []
        for i, r in self.rows("БАНК", 9, None, 32):
            d = as_date(r[10])                     # K — дата
            if not d:
                continue
            expense, income = num(r[18]), num(r[19])   # S расход, T приход
            if not expense and not income:
                continue
            out.append({
                "row": i, "date": d,
                "direction": "income" if income else "expense",
                "amount": income or expense,
                "org_key": txt(r[5]),               # F — код по ИНН
                "org_name": txt(r[1]) or txt(r[15]),  # B отправитель / P корресп.
                "bank": txt(r[9]),                  # J — наименование счёта
                "expense_code": txt(r[6]),          # G
                "cashflow_code": txt(r[7]),         # H
                "doc_no": txt(r[11]), "mfo": txt(r[12]),
                "corr_account": txt(r[14]), "corr_name": txt(r[15]),
                "corr_inn": txt(r[16]) or txt(r[5]),
                "purpose": txt(r[2]), "product_code": txt(r[4]),
                "description": txt(r[17]),
            })
        return out

    def cash_tx(self) -> list[dict]:
        """Лист «КАССА», строки 8.. — все, у которых есть дата."""
        out = []
        for i, r in self.rows("КАССА", 8, None, 28):
            d = as_date(r[2])                      # C — дата
            if not d:
                continue
            income, expense = num(r[10]), num(r[14])   # K приход, O расход
            if not income and not expense:
                continue
            code = txt(r[7])                       # H — код для расходов (2032_М)
            out.append({
                "row": i, "date": d,
                "direction": "income" if income else "expense",
                "amount": income or expense,
                "register": txt(r[0]),             # A — касса
                "org_key": txt(r[3]),              # D — ИНН
                "org_name": txt(r[4]),             # E
                "expense_code": code.split("_")[0],
                "cashflow_code": txt(r[8]),        # I — код платежа
                "purpose": txt(r[9]),              # J — назначение
                "description": txt(r[5]),          # F — наименование платежа
                "division_override": ROW_DIVISION_OVERRIDE.get(("КАССА", i)),
            })
        return out

    def product_openings(self) -> list[dict]:
        """«Остаток ГП»: блоки по объектам (Махстон / Турк / Жби)."""
        return self._stock_blocks("Остаток ГП", code_col=2, name_col=3,
                                  qty_col=5, price_col=6, last=25)

    def material_openings(self) -> list[dict]:
        """«Остаток сырья и запчастей»: блоки по объектам."""
        return self._stock_blocks("Остаток сырья и запчастей", code_col=1, name_col=3,
                                  qty_col=5, price_col=6, last=40)

    def _stock_blocks(self, sheet, code_col, name_col, qty_col, price_col, last):
        out, division = [], ""
        for _i, r in self.rows(sheet, 3, last, 12):
            marker = txt(r[2])
            if marker in ("Махстон", "Турк", "Жби"):
                division = marker
                continue
            name = txt(r[name_col])
            qty, price = num(r[qty_col]), num(r[price_col])
            if not name or name == "0" or not qty:
                continue
            out.append({"division": division, "code": txt(r[code_col]), "name": name,
                        "qty": qty, "price": price})
        return out

    def productions(self) -> list[dict]:
        out = []
        for _i, r in self.rows("Производства Приход ГП", 5, 40, 10):
            division, name = txt(r[2]), txt(r[4])
            qty, cost = num(r[6]), num(r[7])
            if not name or not qty:
                continue
            out.append({"division": division, "code": txt(r[3]), "name": name,
                        "qty": qty, "unit_cost": cost})
        return out

    def receipts(self, sheet: str, kind: str) -> list[dict]:
        """«Приход сырья» / «Приход запчастей» — одна раскладка колонок на оба листа.

        B Месяц, C Дата, D ИНН поставщика, E наименование поставщика, F код,
        G наименование материала, H ед.изм., I дробилка, J вид оплаты,
        K плательщик НДС, L кол-во, M цена без НДС, N сумма без НДС.
        """
        out = []
        for i, r in self.rows(sheet, 7, None, 14):
            d = as_date(r[2])
            qty = num(r[11])
            if not d or not qty:
                continue
            code, name = txt(r[5]), txt(r[6])
            if not name:
                if not code:
                    continue
                # по коду, БЕЗ номера строки: тот же код без названия
                # встречается и в «Расход сырья и запчастей» (спишется же не
                # само по себе) — по номеру строки они не совпали бы, и
                # обе стороны завели бы себе разные материалы на один товар
                name = f"Без наименования (код {code})"
            division = txt(r[8])
            division = DIVISION_ALIASES.get(division, division)
            out.append({
                "date": d, "org_inn": txt(r[3]), "org_name": txt(r[4]),
                "code": code, "name": name, "division": division,
                "payment_type": txt(r[9]), "vat": txt(r[10]).startswith("с учетом"),
                "qty": qty, "price": num(r[12]), "kind": kind,
            })
        return out

    def material_issues(self) -> list[dict]:
        """«Расход сырья и запчастей»: у листа своей полезной даты нет (везде
        стоит дата начала учёта — заглушка), поэтому дату берём фиксированной
        в `run()`, как и для «Производства Приход ГП».

        A дата(не используется), B код сырья, C наим. сырья, E кол-во сырья,
        F код запчасти, G наим. запчасти, I код расходы, J объект,
        M кол-во запчасти, Q сумма UZS (без учёта курса — считаем сами).
        """
        out = []
        for _i, r in self.rows("Расход сырья и запчастей", 6, None, 17):
            raw_code, raw_name = txt(r[1]), txt(r[2])
            spare_code, spare_name = txt(r[5]), txt(r[6])
            if raw_code or raw_name:
                name, kind, qty = raw_name, "raw", num(r[4])
                code = raw_code
            elif spare_code or spare_name:
                name, kind, qty = spare_name, "spare", num(r[12])
                code = spare_code
            else:
                continue
            if not qty:
                continue
            if not name:
                # тот же товар без названия, что и в «Приход» — код должен
                # совпасть с тем, что заведён там (см. receipts())
                if not code:
                    continue
                name = f"Без наименования (код {code})"
            # код расходов в этом листе — с суффиксом объекта («2025_М»),
            # как и в «КАССА»; отделяем его тем же способом (там же и
            # division — свой отдельный столбец, суффикс дублирует его)
            out.append({"name": name, "kind": kind, "division": txt(r[9]),
                        "expense_code": txt(r[8]).split("_")[0], "qty": qty})
        return out

    def sales(self) -> list[dict]:
        """«Продажа Расход ГП»: данные начинаются после сводного блока по строке
        с заголовками (13-я строка).

        B Месяц, C Дата, E ИНН, F клиент, H дробилка (= объект/подразделение),
        I наименование ГП, J объект (стройплощадка клиента — не подразделение),
        K вид оплаты, M кол-во, N цена с НДС, O сумма с НДС, Q сумма НДС(сум),
        S сумма без НДС.
        """
        out = []
        for _i, r in self.rows("Продажа Расход ГП", 14, None, 20):
            name = txt(r[8])
            if not name or name.lower() in ("итого", "всего"):
                continue
            qty = num(r[12])
            d = as_date(r[2])
            if not d or not qty:
                continue
            out.append({
                "date": d, "org_inn": txt(r[4]), "org_name": txt(r[5]),
                "division": txt(r[7]), "site": txt(r[9]), "payment_type": txt(r[10]),
                "name": name, "qty": qty, "price": num(r[13]),
                "vat": num(r[16]) > 0.005,
            })
        return out

    def taxes(self) -> list[dict]:
        """Лист «Налоги»: C долг / D переплата на начало, E начислено,
        F оплачено, G долг / H переплата на конец.

        НДФЛ, ЕСП и ИНПС система начисляет САМА из расчёта зарплаты, а в этой
        книге лист «Зарплата» пуст — начисления вбиты прямо в лист «Налоги».
        Если оставить их как ручные, авто-расчёт всё равно подставит ноль и
        долг на конец «уедет» на сумму начисления. Поэтому у таких налогов
        движение месяца не переносим, а сохраняем ИТОГОВОЕ сальдо — оно и
        попадает в баланс.
        """
        auto_from_payroll = ("ндфл", "есп", "инпс")
        out = []
        for _i, r in self.rows("Налоги", 7, 13, 10):
            name = txt(r[1])
            if not name:
                continue
            start = round(num(r[2]) - num(r[3]), 2)
            end = round(num(r[6]) - num(r[7]), 2)
            accrued, paid = num(r[4]), num(r[5])
            if any(k in name.lower() for k in auto_from_payroll) and accrued:
                start, accrued, paid = end, 0.0, 0.0
            out.append({"name": name, "debt_start": start, "accrued": accrued,
                        "paid": paid, "debt_end": end})
        return out

    def payroll_totals(self) -> list[dict]:
        """Лист «Зарплата  » (с двумя пробелами в имени — НЕ путать с пустым
        листом «Зарплата»): свод начисленной зарплаты по объекту/коду, без
        разбивки по сотрудникам. Отсюда «Расходы <объект>» берёт строку 2012
        «Зарплата производственного персонала» (D8 листа книги) — без нашего
        импорта себестоимость выпуска выходит в разы ниже книжной.

        «Объект» здесь — НАСТОЯЩЕЕ подразделение (включая «АУП» отдельной
        строкой) и именно оно должно попасть в Employee.division: это поле
        читает и себестоимость (см. production.py — там для АУП есть
        отдельное правило), и ведомость «Дт Кт З.п» (начисление кредитуется
        организации «Ойлик(АУП)», отдельной от «Ойлик(Махстон)» — сверено
        с книгой построчно). Код («2012_М») здесь НЕ определяет объект —
        только статью расходов.

        D Объект, E Код платежа, I Начислено (Фин).
        """
        out = []
        for _i, r in self.rows("Зарплата  ", 3, 40, 10):
            division, code, amount = txt(r[3]), txt(r[4]), num(r[8])
            if not division or not amount:
                continue
            out.append({"division": division, "expense_code": code.split("_")[0],
                        "amount": round(amount, 2)})
        return out

    # ---------- справочники книги ----------

    def registry_orgs(self) -> list[dict]:
        """«РЕЕСТР организации»: C ИНН, D наименование, E принадлежит,
        F плательщик НДС, G вид НДС."""
        out, seen = [], set()
        for _i, r in self.rows("РЕЕСТР организации", 6, None, 10):
            name, inn = txt(r[3]), txt(r[2])
            if not name or name == "0":
                continue
            k = (inn.upper(), name.upper())
            if k in seen:
                continue
            seen.add(k)
            out.append({"inn": inn, "name": name, "belongs_to": txt(r[4]) or "Прочие",
                        "nds_payer": txt(r[5]).lower().startswith("да"),
                        "nds_type": txt(r[6])})
        return out

    def nomenclature(self, sheet: str, kind: str) -> list[dict]:
        """«Наименование сырья» / «Наименование запчастей»: B код, E наименование,
        F ед. изм., G местный/импорт."""
        out, seen = [], set()
        for _i, r in self.rows(sheet, 4, None, 10):
            name = txt(r[4])
            if not name or name == "0":
                continue
            if name.upper() in seen:
                continue
            seen.add(name.upper())
            out.append({"code": txt(r[1]), "name": name, "unit": txt(r[5]),
                        "source": txt(r[6]) or "Местный", "warehouse": txt(r[3]),
                        "kind": kind})
        return out

    def gp_nomenclature(self) -> list[dict]:
        """«Наименование ГП»: B код, C наименование, D ед. изм., E краткое."""
        out = []
        for _i, r in self.rows("Наименование ГП", 4, 40, 8):
            name = txt(r[2])
            if not name or name == "0":
                continue
            out.append({"code": txt(r[1]), "name": name, "unit": txt(r[3]),
                        "short_name": txt(r[4])})
        return out

    def services_received(self) -> list[dict]:
        """«Полученные УСЛУГИ» — «Оказанные» в книге пустой (проверено: сумма 0).

        B ИНН, C наименование, D месяц, E дата счёт-фактуры, F вид услуг,
        G код платежа, H назначение, I НДС, J объект,
        K сумма без НДС (сум), M сумма НДС (сум).
        """
        out = []
        for i, r in self.rows("Полученные УСЛУГИ", 6, None, 14):
            d = as_date(r[4])
            net, vat_amt = num(r[10]), num(r[12])
            if not d or (not net and not vat_amt):
                continue
            division = ROW_DIVISION_OVERRIDE.get(("Полученные УСЛУГИ", i)) or txt(r[9])
            out.append({
                "date": d, "org_inn": txt(r[1]), "org_name": txt(r[2]),
                # код платежа — с суффиксом объекта («2027_М»), как в «КАССА»
                # и «Расход сырья и запчастей»; отделяем тем же способом
                "service_type": txt(r[5]), "expense_code": txt(r[6]).split("_")[0],
                "purpose": txt(r[7]), "vat": txt(r[8]).startswith("с учетом"),
                "division": division, "net": net, "vat_amount": vat_amt,
            })
        return out

    def fixed_assets(self) -> float:
        row = next(r for i, r in self.rows("ОС", 2, 2, 10))
        return round(num(row[8]), 2)

    def charter_capital(self) -> float:
        row = next(r for i, r in self.rows("Офис Note", 6, 6, 14))
        return round(num(row[11]), 2)   # L6 — итог «на конец, дебет»

    # Разница план/факт по сырью (лист «С-сть ГП», строка 110/160/170 ОФР)
    # книгой не читаем: приложение считает её заново на лету из документов
    # (Производство, остатки склада, Расход сырья) — см. reports._material_variance.


# ---------------------------------------------------------------- запись

WIPE_MODELS = [
    ProductPrice, Sale, Production, MaterialIssue, MaterialReceipt, Service,
    PayrollEntry, Employee, LoanEntry, Loan, Tax, Transaction, PeriodClose,
    ProductStock, MaterialStock, PeriodSetting,
]


async def wipe(db):
    for model in WIPE_MODELS:
        await db.execute(delete(model))
    # входящие остатки справочников тоже переписываем книгой
    await db.execute(text("UPDATE organizations SET opening_uzs=0, opening_usd=0, "
                          "opening_rate=0, opening_date=NULL, balance_uzs=0, balance_usd=0"))
    await db.execute(text("UPDATE products SET opening_qty=0, opening_cost=0, "
                          "stock_qty=0, avg_cost=0"))
    await db.execute(text("UPDATE materials SET opening_qty=0, opening_cost=0, "
                          "stock_qty=0, avg_cost=0"))
    await db.execute(text("UPDATE bank_accounts SET opening_uzs=0, opening_usd=0, "
                          "opening_date=NULL"))
    await db.execute(text("UPDATE cash_registers SET opening_uzs=0, opening_usd=0, "
                          "opening_date=NULL"))
    await db.commit()
    print("· старые документы удалены")


class Registry:
    """Контрагенты в памяти, ключ — «код по ИНН» из книги.

    Ищем ТОЛЬКО по ключу, без запасного поиска по названию. В книге ключ —
    единственное, по чему БАНК/КАССА связываются с ведомостью, и разные ключи
    с похожим названием («309899765» и «309 899 765 АЗМУР») — это разные
    контрагенты со своими сальдо. Поиск по названию их склеивал, и сальдо
    одного из них пропадало из баланса.
    """

    def __init__(self, db):
        self.db = db
        self.orgs_key: dict[str, Organization] = {}
        self.created_orgs = 0

    async def load(self):
        for o in (await self.db.execute(select(Organization))).scalars().all():
            if o.inn:
                self.orgs_key.setdefault(okey(o.inn).upper(), o)

    def find_org(self, key: str) -> Organization | None:
        return self.orgs_key.get(okey(key).upper()) if key else None

    def remember(self, org: Organization) -> None:
        self.orgs_key[okey(org.inn).upper()] = org

    async def ensure_org(self, key: str, name: str, ledger: str, category: str) -> Organization:
        found = self.find_org(key)
        if found:
            found.ledger = ledger
            found.category = category
            return found
        org = Organization(inn=okey(key), name=(name or key)[:255],
                           ledger=ledger, category=category)
        self.db.add(org)
        await self.db.flush()
        self.remember(org)
        self.created_orgs += 1
        return org


async def run(path: str, do_wipe: bool, rates_path: str | None = None):
    book = Book(path)
    async with AsyncSessionLocal() as db:
        if do_wipe:
            await wipe(db)

        # --- курс доллара ---
        if rates_path:
            with open(rates_path, encoding="utf-8") as f:
                raw = json.load(f)
            rates = {date.fromisoformat(k): float(v) for k, v in raw.items()}
            src = f"файл {rates_path}"
        else:
            rates = book.rates()
            src = "лист «Курс доллара»"
        existing = {d for (d,) in (await db.execute(select(ExchangeRate.rate_date))).all()}
        for d, v in sorted(rates.items()):
            if d in existing:
                await db.execute(
                    text("UPDATE exchange_rates SET rate=:r WHERE rate_date=:d"),
                    {"r": v, "d": d},
                )
            else:
                db.add(ExchangeRate(rate_date=d, rate=v))
        await db.flush()
        # курс, по которому пересчитаны ВСЕ входящие сальдо
        open_rate = rates.get(OPENING_DATE) or 1.0
        print(f"· курс ({src}): {len(rates)} дат, {min(rates.values()):,.2f} … "
              f"{max(rates.values()):,.2f}; на {OPENING_DATE} — {open_rate:,.2f}")

        # --- банковские счета ---
        banks = {}
        have = {(b.name or "").strip(): b for b in
                (await db.execute(select(BankAccount))).scalars().all()}
        for b in book.banks():
            row = have.get(b["name"].strip())
            if row is None:
                row = BankAccount(name=b["name"])
                db.add(row)
            row.account_no, row.mfo = b["account_no"], b["mfo"]
            row.opening_uzs = b["opening"]
            row.opening_usd = round(b["opening"] / open_rate, 2)
            row.opening_date = OPENING_DATE
            await db.flush()
            banks[b["name"].strip()] = row
        print(f"· банковские счета: {len(banks)}, входящий остаток "
              f"{sum(float(x.opening_uzs) for x in banks.values()):,.2f}")

        # --- кассы ---
        tills = {(c.name or "").strip(): c for c in
                 (await db.execute(select(CashRegister))).scalars().all()}
        for name in {t["register"] for t in book.cash_tx() if t["register"]}:
            if name not in tills:
                div = name.replace(" касса", "").strip()
                c = CashRegister(name=name, division=div)
                db.add(c)
                await db.flush()
                tills[name] = c
        main_till = tills.get("Офис касса") or next(iter(tills.values()))
        main_till.opening_uzs = book.cash_opening()
        main_till.opening_usd = round(book.cash_opening() / open_rate, 2)
        main_till.opening_date = OPENING_DATE
        for c in tills.values():
            if c is not main_till:
                c.opening_date = OPENING_DATE
        await db.flush()
        print(f"· кассы: {len(tills)}, входящий остаток {book.cash_opening():,.2f} "
              f"(на «{main_till.name}» — в книге он один общий)")

        # --- справочник контрагентов книги («РЕЕСТР организации») ---
        # Заводим ДО ведомостей: там задаётся только сальдо и вид ведомости, а
        # реквизиты (принадлежность, НДС) есть лишь в реестре.
        reg = Registry(db)
        await reg.load()
        by_name = {(o.name or "").strip().upper(): o
                   for o in (await db.execute(select(Organization))).scalars().all()}
        added = 0
        for row in book.registry_orgs():
            org = reg.find_org(row["inn"]) or by_name.get(row["name"].upper())
            if org is None:
                org = Organization(inn=okey(row["inn"]), name=row["name"][:255])
                db.add(org)
                await db.flush()
                reg.remember(org)
                by_name[org.name.upper()] = org
                added += 1
            org.belongs_to = row["belongs_to"][:120]
            org.nds_payer = row["nds_payer"]
            org.nds_type = row["nds_type"][:60]
        await db.flush()
        print(f"· РЕЕСТР организации: {added} новых")

        # --- номенклатура сырья / запчастей / ГП ---
        mats = {(m.name or "").strip().upper(): m for m in
                (await db.execute(select(Material))).scalars().all()}
        n_mat = 0
        for row in (book.nomenclature("Наименование сырья", "raw")
                    + book.nomenclature("Наименование запчастей", "spare")):
            m = mats.get(row["name"].upper())
            if m is None:
                m = Material(name=row["name"], kind=row["kind"])
                db.add(m)
                await db.flush()
                mats[row["name"].upper()] = m
                n_mat += 1
            m.code, m.unit = row["code"], row["unit"]
            m.source, m.warehouse, m.kind = row["source"], row["warehouse"], row["kind"]

        prods = {(p.name or "").strip().upper(): p for p in
                 (await db.execute(select(Product))).scalars().all()}
        n_prod_dir = 0
        for row in book.gp_nomenclature():
            p = prods.get(row["name"].upper())
            if p is None:
                p = Product(name=row["name"])
                db.add(p)
                await db.flush()
                prods[row["name"].upper()] = p
                n_prod_dir += 1
            p.code, p.unit, p.short_name = row["code"], row["unit"], row["short_name"]
        await db.flush()
        print(f"· номенклатура: сырьё/запчасти +{n_mat} (всего {len(mats)}), "
              f"ГП +{n_prod_dir} (всего {len(prods)})")

        # --- контрагенты: входящее сальдо ---
        opened = 0
        for row in book.ledger_openings():
            org = await reg.ensure_org(row["key"], row["name"], row["ledger"], row["category"])
            org.opening_uzs = row["opening"]
            org.opening_usd = round(row["opening"] / open_rate, 2)
            org.opening_rate = open_rate
            org.opening_date = OPENING_DATE
            opened += 1
        await db.flush()
        print(f"· контрагенты: {opened} строк сальдо, новых организаций {reg.created_orgs}")

        # --- займы ---
        for ln in book.loans():
            db.add(Loan(counterparty=ln["name"], direction=ln["direction"], currency="UZS",
                        opening_uzs=ln["opening"], opening_date=OPENING_DATE,
                        balance=ln["opening"]))
        await db.flush()

        # --- операции БАНК / КАССА ---
        rate_of = {d: v for d, v in rates.items()}
        default_rate = open_rate
        no_org: dict[str, int] = {}
        n_bank = n_cash = 0

        for t in book.bank_tx():
            org = reg.find_org(t["org_key"])
            if org is None and t["org_key"]:
                no_org[t["org_key"]] = no_org.get(t["org_key"], 0) + 1
            r = rate_of.get(t["date"], default_rate) or 1.0
            db.add(Transaction(
                doc_date=t["date"], direction=t["direction"], account="bank", currency="UZS",
                amount=t["amount"], rate=r, amount_uzs=t["amount"],
                amount_usd=round(t["amount"] / r, 2),
                organization_id=org.id if org else None,
                bank_account_id=banks[t["bank"].strip()].id if t["bank"].strip() in banks else None,
                expense_code=t["expense_code"], cashflow_code=t["cashflow_code"],
                doc_no=t["doc_no"], mfo=t["mfo"], corr_account=t["corr_account"],
                corr_name=t["corr_name"][:255], corr_inn=t["corr_inn"][:20],
                purpose=t["purpose"][:200], product_code=t["product_code"][:30],
                category=t["purpose"][:160], description=t["description"],
            ))
            n_bank += 1

        for t in book.cash_tx():
            org = reg.find_org(t["org_key"])
            if org is None and t["org_key"]:
                no_org[t["org_key"]] = no_org.get(t["org_key"], 0) + 1
            till = tills.get(t["register"])
            r = rate_of.get(t["date"], default_rate) or 1.0
            db.add(Transaction(
                doc_date=t["date"], direction=t["direction"], account="kassa", currency="UZS",
                amount=t["amount"], rate=r, amount_uzs=t["amount"],
                amount_usd=round(t["amount"] / r, 2),
                organization_id=org.id if org else None,
                cash_register_id=till.id if till else None,
                cash_register=t["register"][:80],
                division=(t["division_override"] or (till.division if till else ""))[:80],
                expense_code=t["expense_code"], cashflow_code=t["cashflow_code"],
                purpose=t["purpose"][:200], category=t["purpose"][:160],
                description=t["description"],
            ))
            n_cash += 1
        await db.flush()
        print(f"· операции: БАНК {n_bank}, КАССА {n_cash}")
        if no_org:
            top = sorted(no_org.items(), key=lambda x: -x[1])[:10]
            print(f"  не нашли контрагента для {sum(no_org.values())} операций "
                  f"({len(no_org)} ключей), напр.: " + ", ".join(f"{k}×{v}" for k, v in top))

        # --- склады: входящие остатки ---
        for row in book.product_openings():
            p = prods.get(row["name"].upper())
            if p is None:
                p = Product(code=row["code"], name=row["name"], unit="м³")
                db.add(p)
                await db.flush()
                prods[row["name"].upper()] = p
            # opening_cost — это СРЕДНЯЯ ЦЕНА за единицу (так её читает реплей
            # склада в app/stock.py), а не общая сумма остатка
            db.add(ProductStock(product_id=p.id, division=row["division"],
                                opening_qty=row["qty"], opening_cost=row["price"]))
        for row in book.material_openings():
            m = mats.get(row["name"].upper())
            if m is None:
                m = Material(code=row["code"], name=row["name"], kind="raw")
                db.add(m)
                await db.flush()
                mats[row["name"].upper()] = m
            db.add(MaterialStock(material_id=m.id, division=row["division"],
                                 opening_qty=row["qty"], opening_cost=row["price"]))
        await db.flush()

        # --- производство ГП ---
        n_prod = 0
        for row in book.productions():
            p = prods.get(row["name"].upper())
            if p is None:
                continue
            db.add(Production(doc_date=MONTH_START,
                              product_id=p.id, division=row["division"],
                              qty=row["qty"], unit_cost=row["unit_cost"],
                              amount_uzs=round(row["qty"] * row["unit_cost"], 2),
                              note="Книга: Производства Приход ГП"))
            n_prod += 1
        await db.flush()
        print(f"· производство ГП: {n_prod} строк")

        # --- приход сырья / запчастей ---
        n_recv = 0
        no_org_recv: dict[str, int] = {}
        for row in book.receipts("Приход сырья ", "raw") + book.receipts("Приход запчастей", "spare"):
            m = mats.get(row["name"].upper())
            if m is None:
                m = Material(code=row["code"], name=row["name"], kind=row["kind"])
                db.add(m)
                await db.flush()
                mats[row["name"].upper()] = m
            org = reg.find_org(row["org_inn"])
            if org is None and row["org_inn"]:
                no_org_recv[row["org_inn"]] = no_org_recv.get(row["org_inn"], 0) + 1
            db.add(MaterialReceipt(
                doc_date=row["date"], material_id=m.id, organization_id=org.id if org else None,
                division=row["division"][:80], qty=row["qty"], price_uzs=row["price"],
                vat=row["vat"], payment_type=row["payment_type"][:30],
                note=("Книга: Приход сырья" if row["kind"] == "raw" else "Книга: Приход запчастей"),
            ))
            n_recv += 1
        await db.flush()
        print(f"· приход сырья/запчастей: {n_recv} строк")
        if no_org_recv:
            top = sorted(no_org_recv.items(), key=lambda x: -x[1])[:10]
            print(f"  не нашли поставщика для {sum(no_org_recv.values())} приходов "
                  f"({len(no_org_recv)} ключей), напр.: " + ", ".join(f"{k}×{v}" for k, v in top))

        # --- расход сырья / запчастей ---
        # у листа своей полезной даты нет (везде дата начала учёта) — относим
        # весь месячный расход на конец месяца, как и производство ГП
        n_iss = 0
        issue_date = MONTH_END
        for row in book.material_issues():
            m = mats.get(row["name"].upper())
            if m is None:
                m = Material(name=row["name"], kind=row["kind"])
                db.add(m)
                await db.flush()
                mats[row["name"].upper()] = m
            db.add(MaterialIssue(
                doc_date=issue_date, material_id=m.id, division=row["division"][:80],
                expense_code=row["expense_code"][:20], qty=row["qty"],
                note="Книга: Расход сырья и запчастей",
            ))
            n_iss += 1
        await db.flush()
        print(f"· расход сырья/запчастей: {n_iss} строк")

        # --- продажа ГП ---
        n_sale = 0
        no_org_sale: dict[str, int] = {}
        for row in book.sales():
            p = prods.get(row["name"].upper())
            if p is None:
                p = Product(name=row["name"])
                db.add(p)
                await db.flush()
                prods[row["name"].upper()] = p
            org = reg.find_org(row["org_inn"])
            if org is None and row["org_inn"]:
                no_org_sale[row["org_inn"]] = no_org_sale.get(row["org_inn"], 0) + 1
            note = "Книга: Продажа ГП" + (f" ({row['site']})" if row["site"] else "")
            db.add(Sale(
                doc_date=row["date"], product_id=p.id, organization_id=org.id if org else None,
                division=row["division"][:80], qty=row["qty"], price_uzs=row["price"],
                vat=row["vat"], payment_type=row["payment_type"][:30], note=note[:255],
            ))
            n_sale += 1
        await db.flush()
        print(f"· продажа ГП: {n_sale} строк")
        if no_org_sale:
            top = sorted(no_org_sale.items(), key=lambda x: -x[1])[:10]
            print(f"  не нашли клиента для {sum(no_org_sale.values())} продаж "
                  f"({len(no_org_sale)} ключей), напр.: " + ", ".join(f"{k}×{v}" for k, v in top))

        # --- полученные услуги ---
        # net/НДС берём готовыми из книги (не пересчитываем по ставке заново —
        # так сходится с «Сумма всего» листа даже если у части строк НДС нет).
        # recompute_production НЕ вызываем: себестоимость выпуска в этой книге
        # берётся готовой (см. productions()), а не пересчитывается заново.
        n_svc = 0
        no_org_svc: dict[str, int] = {}
        for row in book.services_received():
            org = reg.find_org(row["org_inn"])
            if org is None and row["org_inn"]:
                no_org_svc[row["org_inn"]] = no_org_svc.get(row["org_inn"], 0) + 1
            gross = round(row["net"] + row["vat_amount"], 2)
            note = "Книга: " + row["purpose"] if row["purpose"] else "Книга: Полученные услуги"
            db.add(Service(
                doc_date=row["date"], direction="received", organization_id=org.id if org else None,
                service_type=row["service_type"][:200], expense_code=row["expense_code"][:20],
                division=row["division"][:80], amount=gross, vat=row["vat"],
                net=row["net"], vat_amount=row["vat_amount"], note=note[:255],
            ))
            n_svc += 1
        await db.flush()
        print(f"· полученные услуги: {n_svc} строк")
        if no_org_svc:
            top = sorted(no_org_svc.items(), key=lambda x: -x[1])[:10]
            print(f"  не нашли поставщика услуг для {sum(no_org_svc.values())} строк "
                  f"({len(no_org_svc)} ключей), напр.: " + ", ".join(f"{k}×{v}" for k, v in top))

        # --- налоги ---
        for t in book.taxes():
            db.add(Tax(name=t["name"], period=f"{BOOK_YEAR}-{BOOK_MONTH:02d}",
                       accrued_date=TAX_DATE, debt_start=t["debt_start"],
                       accrued=t["accrued"], paid=t["paid"], debt_end=t["debt_end"]))
        await db.flush()

        # --- зарплата: свод по объекту/коду (лист «Зарплата  ») ---
        # без сотрудников — один «сводный» сотрудник на (объект, код), чтобы
        # попасть в себестоимость выпуска (app/production.py cost_parts)
        # тем же путём, что и книга: Employee.division + Employee.expense_code.
        period = f"{BOOK_YEAR}-{BOOK_MONTH:02d}"
        n_payroll = 0
        for row in book.payroll_totals():
            emp = await db.scalar(
                select(Employee).where(
                    Employee.division == row["division"],
                    Employee.expense_code == row["expense_code"],
                    Employee.full_name == "Начислено (свод, книга)",
                )
            )
            if emp is None:
                emp = Employee(full_name="Начислено (свод, книга)", division=row["division"],
                               expense_code=row["expense_code"], position="—", is_active=False)
                db.add(emp)
                await db.flush()
            db.add(PayrollEntry(employee_id=emp.id, period=period, pay_mode="cash",
                                oklad=row["amount"], gross=row["amount"], net=row["amount"],
                                total_cost=row["amount"]))
            n_payroll += 1
        await db.flush()
        if n_payroll:
            print(f"· зарплата (свод из книги): {n_payroll} строк")

        # --- настройки: ОС и уставный капитал ---
        fa, cap = book.fixed_assets(), book.charter_capital()
        wanted = {
            "fa_cost": fa, "fa_depreciation": 0, "ia_cost": 0, "ia_depreciation": 0,
            "equipment_install": 0, "capital_charter": cap,
            "capital_added": 0, "capital_reserve": 0,
            # дата начала учёта = дата, на которую зафиксированы ВСЕ входящие
            # сальдо (OPENING_DATE). Иначе остатки складов не существуют в
            # колонке «на начало периода» и она выходит пустой.
            "period_start": OPENING_DATE.isoformat(),
        }
        for key, val in wanted.items():
            row = await db.get(Setting, key)
            if row is None:
                db.add(Setting(key=key, value=str(val)))
            else:
                row.value = str(val)
        await db.flush()
        print(f"· ОС {fa:,.2f}; уставный капитал {cap:,.2f}")

        # --- пересчёт ---
        for p in prods.values():
            await recompute_product(db, p.id)
        for m in mats.values():
            await recompute_material(db, m.id)
        await recompute_org_balances(db, None)
        await db.commit()
        print("· пересчёт складов и сальдо выполнен")


if __name__ == "__main__":
    argv = sys.argv[1:]
    rates_file = None
    if "--rates" in argv:
        rates_file = argv[argv.index("--rates") + 1]
        argv = [a for i, a in enumerate(argv)
                if i not in (argv.index("--rates"), argv.index("--rates") + 1)]
    files = [a for a in argv if not a.startswith("--")]
    asyncio.run(run(files[0], "--wipe" in sys.argv, rates_file))
